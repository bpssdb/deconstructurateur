"""
Utilitaires pour la manipulation des fichiers Excel
Version corrigée pour mieux détecter les couleurs
"""

import openpyxl
import xlrd
import tempfile
import os
from typing import Union, List, Tuple, Any, Dict

def num_to_excel_col(n: int) -> str:
    """Convertit un numéro de colonne en lettre Excel"""
    if n <= 0:
        return "?"
    col = ""
    while n > 0:
        n -= 1
        col = chr(ord('A') + n % 26) + col
        n //= 26
    return col

def excel_col_to_num(col_str: str) -> int:
    """Convertit une lettre de colonne Excel en numéro"""
    num = 0
    for char in col_str:
        num = num * 26 + (ord(char.upper()) - ord('A') + 1)
    return num

def load_workbook(file, data_only=False):
    """
    Charge un fichier Excel (.xlsx ou .xls)
    Retourne un workbook openpyxl
    """
    # Déterminer le type de fichier
    filename = file.name.lower()
    
    if filename.endswith('.xlsx'):
        # Fichier .xlsx - utiliser openpyxl directement
        return openpyxl.load_workbook(file, data_only=data_only)
    
    elif filename.endswith('.xls'):
        # Fichier .xls - convertir via xlrd
        return convert_xls_to_openpyxl(file)
    
    else:
        raise ValueError("Format de fichier non supporté. Utilisez .xlsx ou .xls")

def convert_xls_to_openpyxl(file):
    """
    Convertit un fichier .xls en workbook openpyxl
    Préserve les couleurs et le formatage
    """
    # Lire le fichier .xls avec xlrd
    xls_book = xlrd.open_workbook(file_contents=file.read(), formatting_info=True)
    
    # Créer un nouveau workbook openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut
    
    # Obtenir les informations de formatage
    xf_list = xls_book.format_map
    
    # Parcourir toutes les feuilles
    for sheet_idx, sheet_name in enumerate(xls_book.sheet_names()):
        xls_sheet = xls_book.sheet_by_name(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        
        # Copier les données et le formatage
        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell = xls_sheet.cell(row_idx, col_idx)
                
                # Écrire la valeur
                ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell.value)
                
                # Appliquer le formatage si disponible
                if cell.xf_index is not None and cell.xf_index < len(xf_list):
                    xf = xf_list[cell.xf_index]
                    
                    # Extraire la couleur de fond si elle existe
                    if xf.background and hasattr(xf.background, 'pattern_colour_index'):
                        color_idx = xf.background.pattern_colour_index
                        if color_idx and color_idx < len(xls_book.colour_map):
                            rgb = xls_book.colour_map.get(color_idx)
                            if rgb:
                                # Convertir RGB en hex
                                hex_color = '%02x%02x%02x' % rgb[:3]
                                from openpyxl.styles import PatternFill
                                fill = PatternFill(start_color=hex_color, 
                                                 end_color=hex_color, 
                                                 fill_type="solid")
                                ws.cell(row=row_idx + 1, column=col_idx + 1).fill = fill
    
    return wb

def get_sheet_names(workbook) -> List[str]:
    """Retourne la liste des noms de feuilles"""
    return workbook.sheetnames

def get_cell_color(cell) -> Union[str, None]:
    """
    Version améliorée pour extraire la couleur d'une cellule
    Gère mieux les différents formats de couleurs Excel
    """
    try:
        # Vérifier si la cellule a un remplissage
        if not hasattr(cell, 'fill') or not cell.fill:
            return None
        
        fill = cell.fill
        
        # Debug
        # print(f"DEBUG get_cell_color: Cell {cell.coordinate}")
        # print(f"  Fill type: {fill.fill_type}")
        # print(f"  Pattern type: {getattr(fill, 'patternType', 'N/A')}")
        
        # Si pas de remplissage ou remplissage "none"
        if fill.fill_type in [None, 'none', '']:
            return None
        
        # Pour les remplissages de type "solid" (le plus courant)
        if fill.fill_type == 'solid' or (hasattr(fill, 'patternType') and fill.patternType == 'solid'):
            # Essayer d'abord fgColor (couleur de premier plan)
            if hasattr(fill, 'fgColor') and fill.fgColor:
                color = extract_color_value(fill.fgColor)
                if color:
                    return color
            
            # Ensuite start_color
            if hasattr(fill, 'start_color') and fill.start_color:
                color = extract_color_value(fill.start_color)
                if color:
                    return color
        
        # Pour les autres types de patterns, essayer bgColor
        if hasattr(fill, 'bgColor') and fill.bgColor:
            color = extract_color_value(fill.bgColor)
            if color:
                return color
        
        # Dernière tentative : end_color
        if hasattr(fill, 'end_color') and fill.end_color:
            color = extract_color_value(fill.end_color)
            if color:
                return color
                
    except Exception as e:
        print(f"Erreur lors de l'extraction de la couleur pour {cell.coordinate}: {e}")
    
    return None

def extract_color_value(color_obj) -> Union[str, None]:
    """
    Extrait la valeur de couleur d'un objet Color d'openpyxl
    """
    if not color_obj:
        return None
    
    # Debug
    # print(f"  Color object type: {type(color_obj)}")
    # print(f"  Color attributes: {dir(color_obj)}")
    
    # Si c'est une chaîne directe
    if isinstance(color_obj, str):
        return clean_color_hex(color_obj)
    
    # Si l'objet a un attribut rgb
    if hasattr(color_obj, 'rgb') and color_obj.rgb:
        return clean_color_hex(color_obj.rgb)
    
    # Si l'objet a un attribut value
    if hasattr(color_obj, 'value') and color_obj.value:
        return clean_color_hex(str(color_obj.value))
    
    # Si c'est une couleur indexée
    if hasattr(color_obj, 'indexed') and color_obj.indexed is not None:
        # Les couleurs indexées nécessitent une table de correspondance
        # Pour l'instant, on ignore les couleurs indexées
        return None
    
    # Si c'est une couleur de thème
    if hasattr(color_obj, 'theme') and color_obj.theme is not None:
        # Les couleurs de thème nécessitent le thème du document
        # Pour l'instant, on ignore les couleurs de thème
        return None
    
    # Tentative sur la représentation string
    try:
        color_str = str(color_obj)
        if len(color_str) in [6, 8] and all(c in '0123456789ABCDEFabcdef' for c in color_str):
            return clean_color_hex(color_str)
    except:
        pass
    
    return None

def clean_color_hex(color_str: str) -> str:
    """
    Nettoie et normalise une chaîne de couleur hexadécimale
    """
    if not color_str:
        return None
    
    # Si c'est un objet RGB d'openpyxl
    if hasattr(color_str, '__class__') and color_str.__class__.__name__ == 'RGB':
        # Extraire la valeur hex de l'objet RGB
        if hasattr(color_str, 'rgb'):
            color_str = color_str.rgb
        else:
            # Tenter de convertir en string
            color_str = str(color_str)
    
    # Convertir en string si ce n'est pas déjà le cas
    if not isinstance(color_str, str):
        color_str = str(color_str)
    
    # Enlever les espaces et mettre en majuscules
    color_str = color_str.strip().upper()
    
    # Enlever le # s'il est présent
    if color_str.startswith('#'):
        color_str = color_str[1:]
    
    # Si c'est du format ARGB (8 caractères), enlever le canal alpha
    if len(color_str) == 8:
        # Les 2 premiers caractères sont l'alpha
        alpha = color_str[:2]
        color_str = color_str[2:]
        
        # Si alpha est 00 (transparent), ignorer cette couleur
        if alpha == '00':
            return None
    
    # Vérifier que c'est bien un hex valide de 6 caractères
    if len(color_str) == 6 and all(c in '0123456789ABCDEF' for c in color_str):
        # Ignorer le blanc et les couleurs très claires
        if color_str in ['FFFFFF', 'FFFFFE', 'FEFEFE']:
            return None
        return color_str
    
    return None

def rgb_to_hex(rgb: Union[str, Tuple[int, int, int]]) -> str:
    """Convertit RGB en hexadécimal"""
    if isinstance(rgb, str):
        return clean_color_hex(rgb)
    elif isinstance(rgb, (tuple, list)) and len(rgb) >= 3:
        return '%02X%02X%02X' % (int(rgb[0]), int(rgb[1]), int(rgb[2]))
    elif isinstance(rgb, int):
        return f"{rgb:06X}"
    
    return "FFFFFF"

def get_merged_cells_info(worksheet) -> Dict[Tuple[int, int], Dict]:
    """
    Retourne les informations sur les cellules fusionnées
    """
    merged_info = {}
    
    if hasattr(worksheet, 'merged_cells'):
        for merged_range in worksheet.merged_cells.ranges:
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            min_col = merged_range.min_col
            max_col = merged_range.max_col
            
            # Ajouter l'info pour la cellule principale (top-left)
            merged_info[(min_row, min_col)] = {
                'min_row': min_row,
                'max_row': max_row,
                'min_col': min_col,
                'max_col': max_col,
                'span_rows': max_row - min_row + 1,
                'span_cols': max_col - min_col + 1
            }
            
            # Marquer toutes les cellules de la plage comme faisant partie d'une fusion
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    if (row, col) != (min_row, min_col):
                        merged_info[(row, col)] = {
                            'is_merged_cell': True,
                            'main_cell': (min_row, min_col),
                            'min_row': min_row,
                            'max_row': max_row,
                            'min_col': min_col,
                            'max_col': max_col
                        }
    
    return merged_info

def get_cell_value(cell) -> Any:
    """Retourne la valeur d'une cellule de manière sûre"""
    if cell.value is None:
        return ""
    return cell.value

def get_worksheet_dimensions(worksheet) -> Tuple[int, int]:
    """Retourne les dimensions (lignes, colonnes) d'une feuille"""
    return worksheet.max_row, worksheet.max_column