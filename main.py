"""
D√©constructurateur Excel - Application principale
Version avec support des paires de labels altern√©es
"""

import streamlit as st
from datetime import datetime
import json
from typing import List, Dict, Optional
import pandas as pd
from collections import defaultdict

# Import des modules locaux
from utils.excel_utils import load_workbook, get_sheet_names, num_to_excel_col
from utils.color_detector import detect_all_colors
from utils.zone_detector import detect_zones_with_alternating_pairs, detect_zones_with_two_colors
from utils.visualization import create_excel_visualization, create_color_preview_html, create_zone_detail_view, create_dataframe_view
from utils.export import export_to_json
import plotly.express as px

st.set_page_config(
    page_title="üìä D√©constructurateur Excel - 4 Couleurs",
    page_icon="üìä",
    layout="wide"
)

# CSS personnalis√© mis √† jour
st.markdown("""
<style>
    .stDataFrame {
        font-size: 12px;
    }
    .color-preview {
        display: inline-block;
        width: 30px;
        height: 30px;
        border: 2px solid #333;
        border-radius: 4px;
        margin-right: 10px;
        vertical-align: middle;
    }
    .color-section {
        background-color: #f0f0f0;
        padding: 15px;
        border-radius: 10px;
        margin-bottom: 15px;
    }
    .section-header {
        font-weight: bold;
        margin-bottom: 10px;
        font-size: 1.1em;
    }
</style>
""", unsafe_allow_html=True)

def init_session_state():
    """Initialise les variables de session"""
    if 'zones' not in st.session_state:
        st.session_state.zones = []
    if 'current_sheet' not in st.session_state:
        st.session_state.current_sheet = None
    if 'workbook' not in st.session_state:
        st.session_state.workbook = None
    if 'selected_zone' not in st.session_state:
        st.session_state.selected_zone = None
    if 'color_palette' not in st.session_state:
        st.session_state.color_palette = None
    if 'detected_colors' not in st.session_state:
        st.session_state.detected_colors = []
    if 'color_cells' not in st.session_state:
        st.session_state.color_cells = {}
    if 'label_pairs' not in st.session_state:
        st.session_state.label_pairs = []
    if 'all_sheets_zones' not in st.session_state:
        st.session_state.all_sheets_zones = {}
    if 'all_sheets_color_cells' not in st.session_state:
        st.session_state.all_sheets_color_cells = {}

def main():
    """Fonction principale de l'application"""
    init_session_state()
    
    st.title("üìä D√©constructurateur Excel - Syst√®me 4 Couleurs")
    st.markdown("D√©tection automatique des zones avec 2 couleurs de headers horizontaux et 2 couleurs de headers verticaux")
     
    # Upload du fichier
    uploaded_file = st.file_uploader(
        "üìÇ Charger un fichier Excel (.xlsx, .xls)", 
        type=['xlsx', 'xls']
    )
    
    if uploaded_file:
        try:
            # Charger le workbook avec les valeurs calcul√©es
            st.session_state.workbook = load_workbook_with_values(uploaded_file)
            sheet_names = get_sheet_names(st.session_state.workbook)
            
            # √âtape 1: Configuration globale de la palette
            st.header("üé® √âtape 1: Configuration globale des couleurs")
            
            # D√©tection des couleurs sur toutes les feuilles
            if st.button("üîç Analyser les couleurs dans tout le fichier", type="primary"):
                with st.spinner("Analyse des couleurs en cours..."):
                    all_colors = set()
                    color_counts = defaultdict(int)
                    st.session_state.all_sheets_color_cells = {}
                    
                    # Analyser toutes les feuilles
                    for sheet in sheet_names:
                        colors, color_cells = detect_all_colors(
                            st.session_state.workbook, 
                            sheet
                        )
                        
                        # IMPORTANT: Sauvegarder les cellules color√©es par feuille
                        st.session_state.all_sheets_color_cells[sheet] = color_cells
                        
                        # Fusionner les couleurs
                        for color in colors:
                            color_hex = color['hex']
                            color_counts[color_hex] += color['count']
                            all_colors.add(color_hex)
                    
                    # Cr√©er la liste consolid√©e des couleurs
                    consolidated_colors = []
                    for hex_color in all_colors:
                        from utils.color_detector import get_color_name
                        consolidated_colors.append({
                            'hex': hex_color,
                            'name': get_color_name(hex_color),
                            'count': color_counts[hex_color]
                        })
                    
                    # Trier par nombre d'occurrences
                    consolidated_colors.sort(key=lambda x: x['count'], reverse=True)
                    
                    st.session_state.detected_colors = consolidated_colors
                    st.session_state.global_color_analysis = True
                    
                    if len(consolidated_colors) > 0:
                        st.success(f"‚úÖ {len(consolidated_colors)} couleurs uniques d√©tect√©es dans {len(sheet_names)} feuilles!")
                        
                        # Debug : afficher un r√©sum√© des cellules par couleur
                        with st.expander("üîç Debug : D√©tails des couleurs"):
                            for color in consolidated_colors[:5]:
                                st.write(f"**Couleur #{color['hex']} ({color['name']})** : {color['count']} cellules")
                                # Afficher quelques exemples
                                examples = []
                                for sheet, cells_dict in st.session_state.all_sheets_color_cells.items():
                                    if color['hex'] in cells_dict:
                                        for cell in cells_dict[color['hex']][:3]:
                                            examples.append(f"{sheet} - {num_to_excel_col(cell['col'])}{cell['row']}: {cell.get('value', '(vide)')}")
                                        if len(examples) >= 3:
                                            break
                                if examples:
                                    st.write("Exemples :")
                                    for ex in examples[:5]:
                                        st.write(f"  - {ex}")
                    else:
                        st.warning("‚ö†Ô∏è Aucune couleur d√©tect√©e dans le fichier.")
            
            # Afficher les couleurs d√©tect√©es
            if st.session_state.detected_colors and hasattr(st.session_state, 'global_color_analysis'):
                display_detected_colors()
                
                # Configuration de la palette globale
                if not st.session_state.color_palette:
                    configure_color_palette_pairs_global()
                else:
                    display_selected_palette_pairs()
                    
                    # Bouton pour reconfigurer
                    if st.button("üîÑ Reconfigurer la palette"):
                        st.session_state.color_palette = None
                        st.rerun()
            
            # √âtape 2: Traitement des feuilles
            if st.session_state.color_palette:
                st.header("üìÑ √âtape 2: Traitement des feuilles")
                
                # Tabs pour le traitement
                process_tab1, process_tab2 = st.tabs(["üîç Traitement individuel", "‚ö° Traitement global"])
                
                with process_tab1:
                    # S√©lection de la feuille
                    selected_sheet = st.selectbox("üìÑ S√©lectionner une feuille √† traiter", sheet_names)
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("üéØ D√©tecter les zones", key="detect_single"):
                            process_single_sheet(selected_sheet)
                    
                    with col2:
                        # Afficher les zones d√©tect√©es pour cette feuille
                        sheet_zones = st.session_state.get('all_sheets_zones', {}).get(selected_sheet, [])
                        if sheet_zones:
                            st.success(f"‚úÖ {len(sheet_zones)} zones d√©tect√©es")
                    
                    # Affichage des r√©sultats pour la feuille s√©lectionn√©e
                    if selected_sheet in st.session_state.get('all_sheets_zones', {}):
                        st.session_state.zones = st.session_state.all_sheets_zones[selected_sheet]
                        st.session_state.current_sheet = selected_sheet
                        display_sheet_results(selected_sheet)
                
                with process_tab2:
                    st.markdown("### ‚ö° Traitement de toutes les feuilles")
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        if st.button("üöÄ Traiter toutes les feuilles", type="primary"):
                            process_all_sheets(sheet_names)
                    
                    with col2:
                        # Statistiques globales
                        if hasattr(st.session_state, 'all_sheets_zones'):
                            total_zones = sum(len(zones) for zones in st.session_state.all_sheets_zones.values())
                            st.metric("Total zones", total_zones)
                    
                    with col3:
                        # Export global
                        if hasattr(st.session_state, 'all_sheets_zones') and st.session_state.all_sheets_zones:
                            if st.button("üì• Exporter tout en JSON"):
                                json_data = export_all_sheets_json()
                                st.download_button(
                                    label="üíæ T√©l√©charger JSON global",
                                    data=json_data,
                                    file_name=f"excel_complet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                                    mime="application/json"
                                )
                    
                    # Afficher le r√©sum√© par feuille
                    if hasattr(st.session_state, 'all_sheets_zones'):
                        display_global_summary()
                
        except Exception as e:
            st.error(f"‚ùå Erreur lors du chargement du fichier: {str(e)}")
            st.info("Assurez-vous que le fichier n'est pas corrompu et qu'il s'agit bien d'un fichier Excel.")
    
    # Instructions
    display_instructions()

def display_detected_colors():
    """Affiche les couleurs d√©tect√©es avec une visualisation am√©lior√©e"""
    st.subheader("Couleurs trouv√©es dans le fichier:")
    
    if not st.session_state.detected_colors:
        st.warning("Aucune couleur d√©tect√©e dans le fichier.")
        return
    
    # Visualisation de la distribution des couleurs
    st.markdown("### üìä Distribution des couleurs")
    
    if st.session_state.detected_colors:
        # Pr√©parer les donn√©es pour le graphique
        color_data = []
        color_map = {}
        
        # Prendre les 15 couleurs les plus fr√©quentes
        for color in st.session_state.detected_colors[:15]:
            color_name = f"{color['name']} (#{color['hex']})"
            color_data.append({
                'Couleur': color_name,
                'Occurrences': color['count']
            })
            color_map[color_name] = f"#{color['hex']}"
        
        if color_data:
            df_colors = pd.DataFrame(color_data)
            
            fig = px.bar(
                df_colors, 
                x='Couleur', 
                y='Occurrences',
                title=f"Distribution des couleurs (Top {len(color_data)} sur {len(st.session_state.detected_colors)} d√©tect√©es)"
            )
            
            # Appliquer les vraies couleurs aux barres
            colors_list = [color_map.get(name, '#888888') for name in df_colors['Couleur']]
            fig.update_traces(marker_color=colors_list)
            
            # Am√©liorer la mise en page
            fig.update_layout(
                showlegend=False, 
                height=500,
                xaxis_tickangle=-45,
                margin=dict(b=150)  # Plus d'espace en bas pour les labels
            )
            
            st.plotly_chart(fig, use_container_width=True)
            
            # Informations suppl√©mentaires
            total_colored_cells = sum(c['count'] for c in st.session_state.detected_colors)
            st.info(f"üí° Total : {total_colored_cells:,} cellules color√©es d√©tect√©es dans l'ensemble du fichier")

def configure_color_palette_pairs_global():
    """Configure la palette de couleurs globale avec 4 couleurs ind√©pendantes"""
    st.markdown("### üéØ Configuration globale de la palette")
    st.info("Cette palette sera utilis√©e pour toutes les feuilles du fichier Excel")
    
    # Pr√©parer les options de couleurs
    color_options = {
        f"{c['name']} (#{c['hex']})": c['hex'] 
        for c in st.session_state.detected_colors
    }
    
    # Configuration de la couleur des zones
    st.markdown("#### üì¶ 1. Couleur des zones de donn√©es")
    zone_color = st.selectbox(
        "Cellules √† labelliser (donn√©es √† extraire)",
        options=list(color_options.keys()),
        help="Cette couleur sera recherch√©e dans toutes les feuilles"
    )
    
    # Configuration des headers horizontaux
    st.markdown("""
    <div class="color-section">
        <div class="section-header">‚û°Ô∏è 2. Headers Horizontaux (en-t√™tes de colonnes)</div>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    used_colors = [color_options[zone_color]]
    
    with col1:
        available_h1 = [opt for opt in color_options.keys() if color_options[opt] not in used_colors]
        h1_color = st.selectbox(
            "Couleur H1 (premi√®re couleur horizontale)",
            options=available_h1,
            key="global_h1_color",
            help="Premi√®re couleur pour les headers horizontaux"
        )
        if h1_color:
            used_colors.append(color_options[h1_color])
    
    with col2:
        available_h2 = [opt for opt in color_options.keys() if color_options[opt] not in used_colors]
        h2_color = st.selectbox(
            "Couleur H2 (deuxi√®me couleur horizontale)",
            options=available_h2,
            key="global_h2_color",
            help="Deuxi√®me couleur pour les headers horizontaux"
        )
        if h2_color:
            used_colors.append(color_options[h2_color])
    
    # Configuration des headers verticaux
    st.markdown("""
    <div class="color-section">
        <div class="section-header">‚¨áÔ∏è 3. Headers Verticaux (en-t√™tes de lignes)</div>
    </div>
    """, unsafe_allow_html=True)
    
    col3, col4 = st.columns(2)
    
    with col3:
        available_v1 = [opt for opt in color_options.keys() if color_options[opt] not in used_colors]
        v1_color = st.selectbox(
            "Couleur V1 (premi√®re couleur verticale)",
            options=available_v1,
            key="global_v1_color",
            help="Premi√®re couleur pour les headers verticaux"
        )
        if v1_color:
            used_colors.append(color_options[v1_color])
    
    with col4:
        available_v2 = [opt for opt in color_options.keys() if color_options[opt] not in used_colors]
        v2_color = st.selectbox(
            "Couleur V2 (deuxi√®me couleur verticale)",
            options=available_v2,
            key="global_v2_color",
            help="Deuxi√®me couleur pour les headers verticaux"
        )
        if v2_color:
            used_colors.append(color_options[v2_color])
    
    # Explication du syst√®me
    with st.expander("üí° Comment fonctionne le syst√®me √† 4 couleurs ?"):
        st.markdown("""
        **Principe du syst√®me √† 4 couleurs :**
        
        1. **Zone de donn√©es** : Cellules contenant les donn√©es √† extraire (1 couleur)
        
        2. **Headers horizontaux** : 2 couleurs (H1 et H2)
           - En remontant dans une colonne depuis une zone, on collecte tous les headers de la PREMI√àRE couleur trouv√©e
           - On s'arr√™te quand on rencontre l'AUTRE couleur horizontale
        
        3. **Headers verticaux** : 2 couleurs (V1 et V2)
           - En reculant dans une ligne depuis une zone, on collecte tous les headers de la PREMI√àRE couleur trouv√©e
           - On s'arr√™te quand on rencontre l'AUTRE couleur verticale
        
        **Exemple :**
        ```
        [H1] [H1] [H2] [H1]
        [V1] [Z]  [Z]  [Z]
        [V2] [Z]  [Z]  [Z]
        ```
        
        - Cellule Z en B2 : prendra H1 (premi√®re couleur H trouv√©e) et V1 (premi√®re couleur V trouv√©e)
        - Cellule Z en C3 : prendra H2 (premi√®re couleur H trouv√©e) et V2 (premi√®re couleur V trouv√©e)
        """)
    
    # Bouton de validation
    if st.button("‚úÖ Valider la palette globale", type="primary"):
        if h1_color and h2_color and v1_color and v2_color:
            all_colors = [
                color_options[zone_color],
                color_options[h1_color],
                color_options[h2_color],
                color_options[v1_color],
                color_options[v2_color]
            ]
            
            if len(all_colors) == len(set(all_colors)):
                st.session_state.color_palette = {
                    'zone_color': color_options[zone_color],
                    'zone_name': zone_color.split(' (')[0],
                    'h1_color': color_options[h1_color],
                    'h1_name': h1_color.split(' (')[0],
                    'h2_color': color_options[h2_color],
                    'h2_name': h2_color.split(' (')[0],
                    'v1_color': color_options[v1_color],
                    'v1_name': v1_color.split(' (')[0],
                    'v2_color': color_options[v2_color],
                    'v2_name': v2_color.split(' (')[0]
                }
                st.success("‚úÖ Palette globale configur√©e! Vous pouvez maintenant traiter les feuilles.")
                st.rerun()
            else:
                st.error("‚ùå Toutes les couleurs doivent √™tre diff√©rentes !")
        else:
            st.error("‚ùå Veuillez s√©lectionner toutes les couleurs !")
            
def process_single_sheet(sheet_name):
    """Traite une seule feuille avec la palette globale - Version corrig√©e"""
    with st.spinner(f"Traitement de la feuille '{sheet_name}'..."):
        # R√©cup√©rer les cellules color√©es pour cette feuille
        if sheet_name in st.session_state.all_sheets_color_cells:
            color_cells = st.session_state.all_sheets_color_cells[sheet_name]
            st.write(f"üìå Utilisation des couleurs d√©tect√©es pr√©c√©demment")
        else:
            # Si pas encore analys√©, le faire maintenant
            st.warning(f"‚ö†Ô∏è Couleurs non d√©tect√©es pour '{sheet_name}', analyse en cours...")
            colors, color_cells = detect_all_colors(
                st.session_state.workbook, 
                sheet_name
            )
            st.session_state.all_sheets_color_cells[sheet_name] = color_cells
        
        # Debug : v√©rifier que les couleurs sont pr√©sentes
        st.write("**Recherche des couleurs de la palette:**")
        zone_color = st.session_state.color_palette['zone_color']
        zone_cells = color_cells.get(zone_color, [])
        st.write(f"- Zone ({zone_color}): {len(zone_cells)} cellules trouv√©es")
        
        # Afficher les infos pour chaque paire
        if 'label_pairs' in st.session_state.color_palette:
            for i, pair in enumerate(st.session_state.color_palette['label_pairs']):
                h_color = pair['horizontal']['color']
                v_color = pair['vertical']['color']
                h_cells = color_cells.get(h_color, [])
                v_cells = color_cells.get(v_color, [])
                st.write(f"- Paire {i+1}:")
                st.write(f"  - Horizontal ({h_color}): {len(h_cells)} cellules")
                st.write(f"  - Vertical ({v_color}): {len(v_cells)} cellules")
        
        # D√©tecter les zones avec le syst√®me adapt√©
        zones, label_data = detect_zones_with_two_colors(
            st.session_state.workbook,
            sheet_name,
            st.session_state.color_palette,
            color_cells
        )
        
        # Debug : afficher les d√©tails des zones
        if zones:
            total_labels = sum(len(z.get('labels', [])) for z in zones)
            st.write(f"üìä **R√©sultat**: {len(zones)} zones d√©tect√©es, {total_labels} labels trouv√©s")
            
            # Afficher un √©chantillon des labels trouv√©s
            if total_labels > 0:
                with st.expander("Voir un √©chantillon des labels trouv√©s"):
                    for zone in zones[:2]:  # Premi√®res 2 zones
                        if zone.get('labels'):
                            st.write(f"**Zone {zone['id']}:**")
                            # Grouper par type
                            h1_labels = [l for l in zone['labels'] if l['type'] == 'h1']
                            h2_labels = [l for l in zone['labels'] if l['type'] == 'h2']
                            v1_labels = [l for l in zone['labels'] if l['type'] == 'v1']
                            v2_labels = [l for l in zone['labels'] if l['type'] == 'v2']
                            
                            if h1_labels:
                                st.write(f"  H1: {', '.join([l.get('value', '(vide)') for l in h1_labels[:3]])}")
                            if h2_labels:
                                st.write(f"  H2: {', '.join([l.get('value', '(vide)') for l in h2_labels[:3]])}")
                            if v1_labels:
                                st.write(f"  V1: {', '.join([l.get('value', '(vide)') for l in v1_labels[:3]])}")
                            if v2_labels:
                                st.write(f"  V2: {', '.join([l.get('value', '(vide)') for l in v2_labels[:3]])}")
            
            if total_labels == 0:
                st.warning("‚ö†Ô∏è Aucun label trouv√© malgr√© la d√©tection de zones.")
                st.info("V√©rifiez que :")
                st.write("1. Les couleurs des labels sont correctement s√©lectionn√©es")
                st.write("2. Les labels sont positionn√©s directement au-dessus ou √† gauche des zones")
                st.write("3. Il n'y a pas de cellules vides entre les labels et les zones")
        else:
            st.warning("‚ö†Ô∏è Aucune zone d√©tect√©e!")
            st.info("V√©rifiez que la couleur de zone s√©lectionn√©e est bien pr√©sente dans cette feuille.")
        
        # Sauvegarder les zones pour cette feuille
        if 'all_sheets_zones' not in st.session_state:
            st.session_state.all_sheets_zones = {}
        st.session_state.all_sheets_zones[sheet_name] = zones
        
        st.success(f"‚úÖ Traitement termin√© pour '{sheet_name}'!")

def process_all_sheets(sheet_names):
    """Traite toutes les feuilles du fichier"""
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    for idx, sheet_name in enumerate(sheet_names):
        status_text.text(f"Traitement de '{sheet_name}'... ({idx+1}/{len(sheet_names)})")
        process_single_sheet(sheet_name)
        progress_bar.progress((idx + 1) / len(sheet_names))
    
    status_text.text("‚úÖ Traitement termin√©!")
    
    # Afficher le r√©sum√©
    total_zones = sum(len(zones) for zones in st.session_state.all_sheets_zones.values())
    st.success(f"üéâ Traitement termin√©! {total_zones} zones d√©tect√©es dans {len(sheet_names)} feuilles.")

def display_sheet_results(sheet_name):
    """Affiche les r√©sultats pour une feuille sp√©cifique"""
    # Utiliser les fonctions existantes mais avec le contexte de la feuille
    st.header(f"üìä R√©sultats pour '{sheet_name}'")
    
    # Barre d'outils
    tool_col1, tool_col2, tool_col3 = st.columns([1, 1, 1])
    
    with tool_col1:
        if st.button("üîÑ Rafra√Æchir", key=f"refresh_{sheet_name}"):
            st.rerun()
    
    with tool_col2:
        if st.button("üîÄ Fusionner zones proches", key=f"merge_{sheet_name}"):
            from utils.zone_detector import merge_zones
            st.session_state.all_sheets_zones[sheet_name] = merge_zones(
                st.session_state.all_sheets_zones[sheet_name], 
                max_gap=1
            )
            st.success("Zones fusionn√©es!")
            st.rerun()
    
    with tool_col3:
        if st.button("üì• Exporter cette feuille", key=f"export_{sheet_name}"):
            json_data = export_single_sheet_json(sheet_name)
            st.download_button(
                label="üíæ T√©l√©charger",
                data=json_data,
                file_name=f"{sheet_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )
    
    # Affichage avec tabs (r√©utiliser les fonctions existantes)
    tab1, tab2, tab3 = st.tabs(["üìã Vue d'ensemble", "üîç Vue d√©taill√©e", "üìä Statistiques"])
    
    with tab1:
        display_overview_tab_pairs(sheet_name)
    
    with tab2:
        display_detailed_tab_pairs(sheet_name)
    
    with tab3:
        display_statistics_tab_pairs()

def display_global_summary():
    """Affiche un r√©sum√© global de toutes les feuilles trait√©es"""
    st.markdown("### üìä R√©sum√© global")
    
    summary_data = []
    for sheet_name, zones in st.session_state.all_sheets_zones.items():
        total_cells = sum(z['cell_count'] for z in zones)
        total_labels = sum(len(z.get('labels', [])) for z in zones)
        
        summary_data.append({
            'Feuille': sheet_name,
            'Zones': len(zones),
            'Cellules': total_cells,
            'Labels': total_labels
        })
    
    df_summary = pd.DataFrame(summary_data)
    st.dataframe(df_summary, use_container_width=True)
    
    # Graphiques r√©capitulatifs
    col1, col2 = st.columns(2)
    
    with col1:
        fig1 = px.bar(df_summary, x='Feuille', y='Zones', 
                      title="Nombre de zones par feuille")
        st.plotly_chart(fig1, use_container_width=True)
    
    with col2:
        fig2 = px.pie(df_summary, values='Cellules', names='Feuille',
                      title="R√©partition des cellules")
        st.plotly_chart(fig2, use_container_width=True)

def export_single_sheet_json(sheet_name):
    """Exporte les donn√©es d'une seule feuille - Version corrig√©e"""
    zones = st.session_state.all_sheets_zones.get(sheet_name, [])
    return export_to_json_with_four_colors(zones, sheet_name, st.session_state.color_palette)

def export_to_json_with_four_colors(zones, sheet_name, color_palette):
    """Exporte les zones avec le syst√®me √† 4 couleurs en JSON"""
    import json
    from datetime import datetime
    from utils.excel_utils import num_to_excel_col
    
    export_data = {
        "date_export": datetime.now().isoformat(),
        "sheet_name": sheet_name,
        "detection_mode": "four_colors_system",
        "color_palette": {
            "zone_color": f"#{color_palette['zone_color']}",
            "zone_name": color_palette['zone_name']
        }
    }
    
    # Ajouter les couleurs de headers selon le format
    if 'label_pairs' in color_palette:
        export_data["color_palette"]["headers"] = {
            "h1": {
                "color": f"#{color_palette['label_pairs'][0]['horizontal']['color']}",
                "name": color_palette['label_pairs'][0]['horizontal']['name']
            },
            "h2": {
                "color": f"#{color_palette['label_pairs'][1]['horizontal']['color']}" if len(color_palette['label_pairs']) > 1 else "",
                "name": color_palette['label_pairs'][1]['horizontal']['name'] if len(color_palette['label_pairs']) > 1 else ""
            },
            "v1": {
                "color": f"#{color_palette['label_pairs'][0]['vertical']['color']}",
                "name": color_palette['label_pairs'][0]['vertical']['name']
            },
            "v2": {
                "color": f"#{color_palette['label_pairs'][1]['vertical']['color']}" if len(color_palette['label_pairs']) > 1 else "",
                "name": color_palette['label_pairs'][1]['vertical']['name'] if len(color_palette['label_pairs']) > 1 else ""
            }
        }
    else:
        # Format direct
        export_data["color_palette"]["headers"] = {
            "h1": {
                "color": f"#{color_palette.get('h1_color', '')}",
                "name": color_palette.get('h1_name', 'H1')
            },
            "h2": {
                "color": f"#{color_palette.get('h2_color', '')}",
                "name": color_palette.get('h2_name', 'H2')
            },
            "v1": {
                "color": f"#{color_palette.get('v1_color', '')}",
                "name": color_palette.get('v1_name', 'V1')
            },
            "v2": {
                "color": f"#{color_palette.get('v2_color', '')}",
                "name": color_palette.get('v2_name', 'V2')
            }
        }
    
    export_data["zones"] = []
    
    # Exporter les zones
    for zone in zones:
        zone_data = {
            "id": zone['id'],
            "bounds": {
                "min_row": zone['bounds']['min_row'],
                "max_row": zone['bounds']['max_row'],
                "min_col": zone['bounds']['min_col'],
                "max_col": zone['bounds']['max_col'],
                "min_col_letter": num_to_excel_col(zone['bounds']['min_col']),
                "max_col_letter": num_to_excel_col(zone['bounds']['max_col'])
            },
            "cell_count": zone['cell_count'],
            "cells": [],
            "labels": {
                "h1": [],
                "h2": [],
                "v1": [],
                "v2": []
            }
        }
        
        # Exporter les cellules
        for cell in zone['cells']:
            zone_data["cells"].append({
                "address": f"{num_to_excel_col(cell['col'])}{cell['row']}",
                "row": cell['row'],
                "col": cell['col'],
                "col_letter": num_to_excel_col(cell['col']),
                "value": str(cell.get('value', '')) if cell.get('value') is not None else ""
            })
        
        # Organiser les labels par type
        for label in zone.get('labels', []):
            label_type = label.get('type', '')
            if label_type in ['h1', 'h2', 'v1', 'v2']:
                zone_data["labels"][label_type].append({
                    "address": f"{num_to_excel_col(label['col'])}{label['row']}",
                    "row": label['row'],
                    "col": label['col'],
                    "col_letter": num_to_excel_col(label['col']),
                    "value": str(label.get('value', '')) if label.get('value') is not None else "",
                    "distance": label.get('distance', 0)
                })
        
        export_data["zones"].append(zone_data)
    
    return json.dumps(export_data, indent=2, ensure_ascii=False)

def export_all_sheets_json():
    """Exporte toutes les feuilles dans un format global"""
    export_data = {
        "date_export": datetime.now().strftime("%Y-%m-%d"),
        "color_palette": {
            "zone_color": f"#{st.session_state.color_palette['zone_color']}",
            "zone_name": st.session_state.color_palette['zone_name']
        },
        "tags": []
    }
    
    # Ajouter la configuration des headers
    if 'label_pairs' in st.session_state.color_palette:
        export_data["color_palette"]["headers"] = {
            "h1": {
                "color": f"#{st.session_state.color_palette['label_pairs'][0]['horizontal']['color']}",
                "name": st.session_state.color_palette['label_pairs'][0]['horizontal']['name']
            },
            "h2": {
                "color": f"#{st.session_state.color_palette['label_pairs'][1]['horizontal']['color']}" if len(st.session_state.color_palette['label_pairs']) > 1 else "",
                "name": st.session_state.color_palette['label_pairs'][1]['horizontal']['name'] if len(st.session_state.color_palette['label_pairs']) > 1 else ""
            },
            "v1": {
                "color": f"#{st.session_state.color_palette['label_pairs'][0]['vertical']['color']}",
                "name": st.session_state.color_palette['label_pairs'][0]['vertical']['name']
            },
            "v2": {
                "color": f"#{st.session_state.color_palette['label_pairs'][1]['vertical']['color']}" if len(st.session_state.color_palette['label_pairs']) > 1 else "",
                "name": st.session_state.color_palette['label_pairs'][1]['vertical']['name'] if len(st.session_state.color_palette['label_pairs']) > 1 else ""
            }
        }
    
    # Parcourir toutes les feuilles et cr√©er les tags
    tag_id = 1
    for sheet_name, zones in st.session_state.all_sheets_zones.items():
        for zone in zones:
            for cell in zone['cells']:
                # Cr√©er un tag pour chaque cellule de zone
                labels = [sheet_name]  # Le nom de la feuille est toujours le premier label
                source_cells = []
                
                # Collecter les labels H1
                h1_labels = [l for l in zone.get('labels', []) if l.get('type') == 'h1']
                for label in h1_labels:
                    if label.get('value'):
                        labels.append(f"H1:{label['value']}")
                        source_cells.append(f"{num_to_excel_col(label['col'])}{label['row']}")
                
                # Collecter les labels H2
                h2_labels = [l for l in zone.get('labels', []) if l.get('type') == 'h2']
                for label in h2_labels:
                    if label.get('value'):
                        labels.append(f"H2:{label['value']}")
                        source_cells.append(f"{num_to_excel_col(label['col'])}{label['row']}")
                
                # Collecter les labels V1
                v1_labels = [l for l in zone.get('labels', []) if l.get('type') == 'v1']
                for label in v1_labels:
                    if label.get('value'):
                        labels.append(f"V1:{label['value']}")
                        source_cells.append(f"{num_to_excel_col(label['col'])}{label['row']}")
                
                # Collecter les labels V2
                v2_labels = [l for l in zone.get('labels', []) if l.get('type') == 'v2']
                for label in v2_labels:
                    if label.get('value'):
                        labels.append(f"V2:{label['value']}")
                        source_cells.append(f"{num_to_excel_col(label['col'])}{label['row']}")
                
                # Ajouter la cellule elle-m√™me
                source_cells.append(f"{num_to_excel_col(cell['col'])}{cell['row']}")
                
                tag = {
                    "id": tag_id,
                    "sheet_name": sheet_name,
                    "row": cell['row'],
                    "col": cell['col'],
                    "cell_address": f"{num_to_excel_col(cell['col'])}{cell['row']}",
                    "value": str(cell.get('value', '')),
                    "labels": labels,
                    "source_cells": source_cells,
                    "zone_id": zone['id']
                }
                
                export_data["tags"].append(tag)
                tag_id += 1
    
    return json.dumps(export_data, indent=2, ensure_ascii=False)


    """Configure la palette de couleurs avec syst√®me de paires"""
    st.header("üéØ √âtape 2: Configuration de la palette avec paires altern√©es")
    
    # Pr√©parer les options de couleurs
    color_options = {
        f"{c['name']} (#{c['hex']})": c['hex'] 
        for c in st.session_state.detected_colors
    }
    
    # Configuration de la couleur des zones
    st.markdown("### üì¶ 1. Couleur des zones de donn√©es")
    zone_color = st.selectbox(
        "Cellules √† labelliser (donn√©es √† compl√©ter par le LLM)",
        options=list(color_options.keys()),
        help="S√©lectionnez la couleur des cellules qui contiennent les donn√©es √† traiter"
    )
    
    # Configuration des paires de labels
    st.markdown("### üè∑Ô∏è 2. Paires de labels (en-t√™tes altern√©s)")
    
    # Nombre de paires
    num_pairs = st.number_input("Nombre de paires de labels", min_value=1, max_value=5, value=2)
    
    # Configuration de chaque paire
    pairs = []
    used_colors = [color_options[zone_color]]  # La couleur de zone est d√©j√† utilis√©e
    
    for i in range(num_pairs):
        st.markdown(f"""
        <div class="pair-container">
            <div class="pair-header">üîó Paire {i+1}</div>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Filtrer les options pour √©viter les doublons
            available_h = [opt for opt in color_options.keys() if color_options[opt] not in used_colors]
            
            h_color = st.selectbox(
                f"Couleur horizontale (colonnes)",
                options=available_h,
                key=f"h_color_{i}",
                help=f"Labels horizontaux pour la paire {i+1}"
            )
            if h_color:
                used_colors.append(color_options[h_color])
        
        with col2:
            # Filtrer les options pour √©viter les doublons
            available_v = [opt for opt in color_options.keys() if color_options[opt] not in used_colors]
            
            v_color = st.selectbox(
                f"Couleur verticale (lignes)",
                options=available_v,
                key=f"v_color_{i}",
                help=f"Labels verticaux pour la paire {i+1}"
            )
            if v_color:
                used_colors.append(color_options[v_color])
        
        if h_color and v_color:
            pairs.append({
                'horizontal': {
                    'color': color_options[h_color],
                    'name': f"Headers H{i+1} ({h_color.split(' (')[0]})"
                },
                'vertical': {
                    'color': color_options[v_color],
                    'name': f"Headers V{i+1} ({v_color.split(' (')[0]})"
                }
            })
    
    # Explication du syst√®me d'alternance
    with st.expander("üí° Comment fonctionne le syst√®me de paires altern√©es ?"):
        st.markdown("""
        **Principe des paires altern√©es :**
        
        1. **Zones de donn√©es** : Les cellules de la couleur s√©lectionn√©e qui contiennent les donn√©es √† traiter
        
        2. **Paires de labels** : Chaque paire contient :
           - Une couleur pour les labels **horizontaux** (en-t√™tes de colonnes)
           - Une couleur pour les labels **verticaux** (en-t√™tes de lignes)
        
        3. **Logique d'alternance** :
           - En remontant dans une colonne, on collecte TOUS les labels horizontaux jusqu'√† rencontrer un label vertical de la M√äME paire
           - En reculant dans une ligne, on collecte TOUS les labels verticaux jusqu'√† rencontrer un label horizontal de la M√äME paire
           - Cela permet de g√©rer des structures complexes avec plusieurs niveaux de headers
        
        **Exemple concret :**
        ```
        [H1] [H1] [H1]  <- Paire 1 Horizontal
        [V1] [Z]  [Z]   <- V1: Paire 1 Vertical, Z: Zone de donn√©es
        [V1] [Z]  [Z]
        ```
        
        Dans cet exemple, chaque cellule Z aura comme labels :
        - Le H1 au-dessus (s'arr√™te car pas de V1 entre les deux)
        - Le V1 √† gauche (s'arr√™te car pas de H1 entre les deux)
        """)
    
    # Bouton de validation
    if st.button("‚úÖ Valider et d√©tecter les zones", type="primary"):
        if len(pairs) == num_pairs and all(p['horizontal']['color'] != p['vertical']['color'] for p in pairs):
            # V√©rifier que toutes les couleurs sont uniques
            all_colors = [color_options[zone_color]]
            for p in pairs:
                all_colors.extend([p['horizontal']['color'], p['vertical']['color']])
            
            if len(all_colors) == len(set(all_colors)):
                validate_and_detect_zones_pairs(
                    selected_sheet, 
                    color_options[zone_color],
                    zone_color.split(' (')[0],
                    pairs
                )
            else:
                st.error("‚ùå Toutes les couleurs doivent √™tre diff√©rentes !")
        else:
            st.error("‚ùå Veuillez configurer toutes les paires avec des couleurs diff√©rentes !")
    
    # Afficher la palette s√©lectionn√©e
    if st.session_state.color_palette:
        display_selected_palette_pairs()

def validate_and_detect_zones_pairs(selected_sheet, zone_color, zone_name, pairs):
    """Valide la palette et lance la d√©tection des zones avec paires altern√©es"""
    st.session_state.color_palette = {
        'zone_color': zone_color,
        'zone_name': zone_name,
        'label_pairs': pairs
    }
    st.session_state.label_pairs = pairs
    
    # D√©tecter les zones
    with st.spinner("D√©tection des zones avec logique de paires altern√©es..."):
        zones, label_data = detect_zones_with_alternating_pairs(
            st.session_state.workbook,
            selected_sheet,
            st.session_state.color_palette,
            st.session_state.color_cells
        )
        st.session_state.zones = zones
        st.session_state.label_data = label_data
        st.success(f"‚úÖ {len(zones)} zones d√©tect√©es avec leurs labels altern√©s!")

def display_selected_palette_pairs():
    """Affiche la palette s√©lectionn√©e - Version 4 couleurs"""
    st.subheader("Palette configur√©e:")
    
    # Zone de donn√©es
    st.markdown(f"""
    <div style="display: flex; align-items: center; margin: 10px 0;">
        <div class="color-preview" style="background-color: #{st.session_state.color_palette['zone_color']}; margin-right: 10px;"></div>
        <strong>Zones de donn√©es:</strong> {st.session_state.color_palette['zone_name']}
    </div>
    """, unsafe_allow_html=True)
    
    # Headers horizontaux
    st.markdown("**Headers Horizontaux:**")
    col1, col2 = st.columns(2)
    
    with col1:
        if 'h1_color' in st.session_state.color_palette:
            st.markdown(f"""
            <div style="display: flex; align-items: center; margin: 10px 0;">
                <div class="color-preview" style="background-color: #{st.session_state.color_palette['h1_color']}; width: 25px; height: 25px;"></div>
                <span>H1: {st.session_state.color_palette['h1_name']}</span>
            </div>
            """, unsafe_allow_html=True)
    
    with col2:
        if 'h2_color' in st.session_state.color_palette:
            st.markdown(f"""
            <div style="display: flex; align-items: center; margin: 10px 0;">
                <div class="color-preview" style="background-color: #{st.session_state.color_palette['h2_color']}; width: 25px; height: 25px;"></div>
                <span>H2: {st.session_state.color_palette['h2_name']}</span>
            </div>
            """, unsafe_allow_html=True)
    
    # Headers verticaux
    st.markdown("**Headers Verticaux:**")
    col3, col4 = st.columns(2)
    
    with col3:
        if 'v1_color' in st.session_state.color_palette:
            st.markdown(f"""
            <div style="display: flex; align-items: center; margin: 10px 0;">
                <div class="color-preview" style="background-color: #{st.session_state.color_palette['v1_color']}; width: 25px; height: 25px;"></div>
                <span>V1: {st.session_state.color_palette['v1_name']}</span>
            </div>
            """, unsafe_allow_html=True)
    
    with col4:
        if 'v2_color' in st.session_state.color_palette:
            st.markdown(f"""
            <div style="display: flex; align-items: center; margin: 10px 0;">
                <div class="color-preview" style="background-color: #{st.session_state.color_palette['v2_color']}; width: 25px; height: 25px;"></div>
                <span>V2: {st.session_state.color_palette['v2_name']}</span>
            </div>
            """, unsafe_allow_html=True)


def display_results(selected_sheet):
    """Affiche les r√©sultats de la d√©tection avec visualisation adapt√©e aux paires"""
    st.header("üìä Visualisation et √©dition")
    
    # Barre d'outils
    tool_col1, tool_col2, tool_col3, tool_col4 = st.columns([1, 1, 1, 1])
    
    with tool_col1:
        if st.button("üîÑ Rafra√Æchir la vue"):
            st.rerun()
    
    with tool_col2:
        if st.button("üîÄ Fusionner zones proches"):
            from utils.zone_detector import merge_zones
            st.session_state.zones = merge_zones(st.session_state.zones, max_gap=1)
            st.success("Zones fusionn√©es!")
            st.rerun()
    
    with tool_col3:
        if st.button("‚ûï Nouvelle zone manuelle"):
            st.session_state.show_manual_zone = True
    
    with tool_col4:
        if st.button("üì• T√©l√©charger JSON"):
            json_data = export_to_json_pairs(
                st.session_state.zones,
                st.session_state.current_sheet,
                st.session_state.color_palette
            )
            st.download_button(
                label="üíæ T√©l√©charger",
                data=json_data,
                file_name=f"zones_pairs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )
    
    # Affichage principal avec tabs
    tab1, tab2, tab3 = st.tabs(["üìã Vue d'ensemble", "üîç Vue d√©taill√©e", "üìä Statistiques"])
    
    with tab1:
        display_overview_tab_pairs(selected_sheet)
    
    with tab2:
        display_detailed_tab_pairs(selected_sheet)
    
    with tab3:
        display_statistics_tab_pairs()

def display_overview_tab_pairs(selected_sheet):
    """Affiche l'onglet vue d'ensemble - Version 4 couleurs"""
    # Sous-tabs pour diff√©rentes vues
    view_tab1, view_tab2 = st.tabs(["üó∫Ô∏è Vue sch√©matique", "üìã Vue tableau"])
    
    with view_tab1:
        col1, col2 = st.columns([3, 1])
        
        with col1:
            # Vue Plotly adapt√©e pour 4 couleurs
            adapted_palette = st.session_state.color_palette.copy()
            adapted_palette['label_colors'] = {
                'h1': {'color': st.session_state.color_palette['h1_color'], 'name': st.session_state.color_palette['h1_name']},
                'h2': {'color': st.session_state.color_palette['h2_color'], 'name': st.session_state.color_palette['h2_name']},
                'v1': {'color': st.session_state.color_palette['v1_color'], 'name': st.session_state.color_palette['v1_name']},
                'v2': {'color': st.session_state.color_palette['v2_color'], 'name': st.session_state.color_palette['v2_name']}
            }
            
            fig = create_excel_visualization(
                st.session_state.workbook,
                selected_sheet,
                st.session_state.zones,
                st.session_state.selected_zone,
                adapted_palette
            )
            st.plotly_chart(fig, use_container_width=True)
            
            # L√©gende pour 4 couleurs
            st.markdown("### üéØ L√©gende")
            
            # Zone
            st.markdown(f"""
            <div style="display: flex; align-items: center; margin: 5px 0;">
                <div style="width: 20px; height: 20px; background-color: #{st.session_state.color_palette['zone_color']}; border: 1px solid black; margin-right: 10px;"></div>
                <span>Zones de donn√©es</span>
            </div>
            """, unsafe_allow_html=True)
            
            # Headers
            col_leg1, col_leg2 = st.columns(2)
            
            with col_leg1:
                st.markdown("**Headers Horizontaux:**")
                st.markdown(f"""
                <div style="display: flex; align-items: center; margin: 5px 0;">
                    <div style="width: 15px; height: 15px; background-color: #{st.session_state.color_palette['h1_color']}; border: 1px solid black; margin-right: 5px;"></div>
                    <span style="font-size: 0.9em;">H1: {st.session_state.color_palette['h1_name']}</span>
                </div>
                <div style="display: flex; align-items: center; margin: 5px 0;">
                    <div style="width: 15px; height: 15px; background-color: #{st.session_state.color_palette['h2_color']}; border: 1px solid black; margin-right: 5px;"></div>
                    <span style="font-size: 0.9em;">H2: {st.session_state.color_palette['h2_name']}</span>
                </div>
                """, unsafe_allow_html=True)
            
            with col_leg2:
                st.markdown("**Headers Verticaux:**")
                st.markdown(f"""
                <div style="display: flex; align-items: center; margin: 5px 0;">
                    <div style="width: 15px; height: 15px; background-color: #{st.session_state.color_palette['v1_color']}; border: 1px solid black; margin-right: 5px;"></div>
                    <span style="font-size: 0.9em;">V1: {st.session_state.color_palette['v1_name']}</span>
                </div>
                <div style="display: flex; align-items: center; margin: 5px 0;">
                    <div style="width: 15px; height: 15px; background-color: #{st.session_state.color_palette['v2_color']}; border: 1px solid black; margin-right: 5px;"></div>
                    <span style="font-size: 0.9em;">V2: {st.session_state.color_palette['v2_name']}</span>
                </div>
                """, unsafe_allow_html=True)
        
        with col2:
            st.markdown("### üéÆ Contr√¥les rapides")
            
            # Liste des zones avec statistiques de labels
            for zone in st.session_state.zones:
                # Compter les labels par type
                label_counts = {'h1': 0, 'h2': 0, 'v1': 0, 'v2': 0}
                for label in zone.get('labels', []):
                    label_type = label.get('type', '')
                    if label_type in label_counts:
                        label_counts[label_type] += 1
                
                # Afficher la zone
                with st.container():
                    if st.button(f"Zone {zone['id']} ({zone['cell_count']} cellules)", 
                                key=f"select_zone_{zone['id']}",
                                use_container_width=True):
                        st.session_state.selected_zone = zone['id']
                        st.rerun()
                    
                    # Afficher les stats de labels
                    stats_text = f"H1:{label_counts['h1']} H2:{label_counts['h2']} V1:{label_counts['v1']} V2:{label_counts['v2']}"
                    st.caption(stats_text)
    
    with view_tab2:
        st.markdown("### üìä Vue tableau avec contenu des cellules")
        
        # Options d'affichage
        col1, col2, col3 = st.columns([1, 1, 2])
        with col1:
            show_colors = st.checkbox("Afficher les couleurs", value=True)
        with col2:
            max_rows = st.number_input("Nombre de lignes max", min_value=10, max_value=200, value=50)
        
        # Cr√©er la vue tableau
        try:
            adapted_palette = st.session_state.color_palette.copy()
            adapted_palette['label_colors'] = {
                'h1': {'color': st.session_state.color_palette['h1_color']},
                'h2': {'color': st.session_state.color_palette['h2_color']},
                'v1': {'color': st.session_state.color_palette['v1_color']},
                'v2': {'color': st.session_state.color_palette['v2_color']}
            }
            
            df_styled = create_dataframe_view(
                st.session_state.workbook,
                selected_sheet,
                st.session_state.zones if show_colors else None,
                adapted_palette if show_colors else None,
                max_rows=max_rows
            )
            
            if show_colors and hasattr(df_styled, 'to_html'):
                st.markdown(df_styled.to_html(), unsafe_allow_html=True)
            else:
                st.dataframe(df_styled, use_container_width=True, height=600)
                
        except Exception as e:
            st.error(f"Erreur lors de la cr√©ation de la vue tableau: {str(e)}")

def display_detailed_tab_pairs(selected_sheet):
    """Affiche l'onglet vue d√©taill√©e pour les paires"""
    if not st.session_state.selected_zone:
        st.info("üëÜ S√©lectionnez une zone dans l'onglet 'Vue d'ensemble' pour voir les d√©tails")
        return
    
    zone = next((z for z in st.session_state.zones if z['id'] == st.session_state.selected_zone), None)
    if not zone:
        return
    
    # En-t√™te avec navigation
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        st.markdown(f"### Zone {zone['id']}")
    with col2:
        if st.button("‚¨ÖÔ∏è Zone pr√©c√©dente", disabled=zone['id'] == 1):
            st.session_state.selected_zone = max(1, zone['id'] - 1)
            st.rerun()
    with col3:
        if st.button("Zone suivante ‚û°Ô∏è", disabled=zone['id'] == len(st.session_state.zones)):
            st.session_state.selected_zone = min(len(st.session_state.zones), zone['id'] + 1)
            st.rerun()
    
    # TABS pour diff√©rentes vues de la zone
    detail_view_tab1, detail_view_tab2, detail_view_tab3 = st.tabs([
        "üó∫Ô∏è Vue sch√©matique", 
        "üìã Vue tableau", 
        "üìä Labels par paire"
    ])
    
    with detail_view_tab1:
        # Vue sch√©matique (Plotly)
        st.markdown("#### üîç Vue zoom√©e de la zone")
        zoom_fig = create_zone_detail_view_pairs(
            st.session_state.workbook,
            selected_sheet,
            zone,
            st.session_state.color_palette
        )
        st.plotly_chart(zoom_fig, use_container_width=True)
    
    with detail_view_tab2:
        # Vue tableau d√©taill√©e
        st.markdown("#### üìã Vue tableau de la zone")
        
        # Options d'affichage
        col1, col2, col3 = st.columns(3)
        with col1:
            show_markers = st.checkbox("Afficher les marqueurs visuels", value=True)
        with col2:
            table_style = st.selectbox("Style du tableau", ["Standard", "Avec marqueurs"], index=1)
        with col3:
            show_legend = st.checkbox("Afficher la l√©gende", value=True)
        
        try:
            if table_style == "Avec marqueurs":
                styled_table = create_zone_detail_table_view_pairs_enhanced(
                    st.session_state.workbook,
                    selected_sheet,
                    zone,
                    st.session_state.color_palette,
                    show_markers
                )
            else:
                styled_table = create_zone_detail_table_view_pairs(
                    st.session_state.workbook,
                    selected_sheet,
                    zone,
                    st.session_state.color_palette
                )
            
            # Afficher le tableau
            if hasattr(styled_table, 'to_html'):
                st.markdown(styled_table.to_html(escape=False), unsafe_allow_html=True)
            else:
                st.dataframe(styled_table, use_container_width=True)
            
            # L√©gende
            if show_legend:
                st.markdown("#### üé® L√©gende")
                
                # Zone
                zone_color = st.session_state.color_palette['zone_color']
                st.markdown(f"""
                <div style="display: flex; align-items: center; margin: 5px 0;">
                    <div style="width: 20px; height: 20px; background-color: #{zone_color}; opacity: 0.3; border: 3px solid #{zone_color}; margin-right: 10px;"></div>
                    <span>üîµ Cellules de zone</span>
                </div>
                """, unsafe_allow_html=True)
                
                # Paires
                if 'label_pairs' in st.session_state.color_palette:
                    for i, pair in enumerate(st.session_state.color_palette['label_pairs']):
                        st.markdown(f"**Paire {i+1}:**")
                        leg_col1, leg_col2 = st.columns(2)
                        
                        with leg_col1:
                            h_color = pair['horizontal']['color']
                            st.markdown(f"""
                            <div style="display: flex; align-items: center; margin: 5px 0;">
                                <div style="width: 20px; height: 20px; background-color: #{h_color}; opacity: 0.5; border: 2px solid #{h_color}; margin-right: 10px;"></div>
                                <span>‚û°Ô∏è {pair['horizontal']['name']}</span>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        with leg_col2:
                            v_color = pair['vertical']['color']
                            st.markdown(f"""
                            <div style="display: flex; align-items: center; margin: 5px 0;">
                                <div style="width: 20px; height: 20px; background-color: #{v_color}; opacity: 0.5; border: 2px solid #{v_color}; margin-right: 10px;"></div>
                                <span>‚¨áÔ∏è {pair['vertical']['name']}</span>
                            </div>
                            """, unsafe_allow_html=True)
        
        except Exception as e:
            st.error(f"Erreur lors de la cr√©ation de la vue tableau: {str(e)}")
            st.info("Essayez de r√©duire la taille de la zone ou v√©rifiez vos donn√©es.")
    
    with detail_view_tab3:
        # Tableau des labels avec regroupement par paire
        st.markdown("#### üìä Labels identifi√©s (par paire)")
        
        if zone.get('labels'):
            # Regrouper les labels par paire et direction
            labels_by_pair = defaultdict(lambda: {'horizontal': [], 'vertical': []})
            
            for label in zone['labels']:
                if 'pair_id' in label:
                    labels_by_pair[label['pair_id']][label['direction']].append(label)
            
            # Afficher chaque paire
            for pair_id in sorted(labels_by_pair.keys()):
                pair_labels = labels_by_pair[pair_id]
                pair_name = f"Paire {pair_id + 1}"
                
                if pair_id < len(st.session_state.color_palette.get('label_pairs', [])):
                    pair_config = st.session_state.color_palette['label_pairs'][pair_id]
                
                with st.expander(f"üîó {pair_name} ({len(pair_labels['horizontal'])} H, {len(pair_labels['vertical'])} V)"):
                    col1, col2 = st.columns(2)
                    
                    # Labels horizontaux
                    with col1:
                        st.markdown("**Labels Horizontaux**")
                        if pair_labels['horizontal']:
                            h_data = []
                            for label in pair_labels['horizontal']:
                                from utils.excel_utils import num_to_excel_col
                                h_data.append({
                                    'Position': f"{num_to_excel_col(label['col'])}{label['row']}",
                                    'Valeur': label.get('value', ''),
                                    'Distance': label.get('distance', 0)
                                })
                            st.dataframe(pd.DataFrame(h_data), use_container_width=True)
                        else:
                            st.info("Aucun label horizontal")
                    
                    # Labels verticaux
                    with col2:
                        st.markdown("**Labels Verticaux**")
                        if pair_labels['vertical']:
                            v_data = []
                            for label in pair_labels['vertical']:
                                from utils.excel_utils import num_to_excel_col
                                v_data.append({
                                    'Position': f"{num_to_excel_col(label['col'])}{label['row']}",
                                    'Valeur': label.get('value', ''),
                                    'Distance': label.get('distance', 0)
                                })
                            st.dataframe(pd.DataFrame(v_data), use_container_width=True)
                        else:
                            st.info("Aucun label vertical")
        else:
            st.warning("Aucun label identifi√© pour cette zone")
    
    # Informations de la zone
    st.markdown("#### üìç Informations de la zone")
    info_col1, info_col2 = st.columns(2)
    
    with info_col1:
        from utils.excel_utils import num_to_excel_col
        st.write(f"**Lignes:** {zone['bounds']['min_row']} √† {zone['bounds']['max_row']}")
        st.write(f"**Colonnes:** {num_to_excel_col(zone['bounds']['min_col'])} √† {num_to_excel_col(zone['bounds']['max_col'])}")
        st.write(f"**Nombre de cellules:** {zone['cell_count']}")
    
    with info_col2:
        st.write("**√âchantillon de valeurs:**")
        sample_values = []
        for cell in zone['cells'][:5]:
            if cell.get('value'):
                sample_values.append(f"‚Ä¢ {cell['value']}")
        if sample_values:
            st.write("\n".join(sample_values))
        else:
            st.write("(cellules vides)")

def display_statistics_tab_pairs():
    """Affiche les statistiques - Version 4 couleurs"""
    if not st.session_state.zones:
        st.info("Aucune zone d√©tect√©e pour afficher les statistiques")
        return
    
    # Statistiques globales
    col1, col2, col3, col4 = st.columns(4)
    
    total_zones = len(st.session_state.zones)
    total_cells = sum(z['cell_count'] for z in st.session_state.zones)
    total_labels = sum(len(z.get('labels', [])) for z in st.session_state.zones)
    
    col1.metric("üì¶ Zones", total_zones)
    col2.metric("üìã Cellules totales", total_cells)
    col3.metric("üè∑Ô∏è Labels totaux", total_labels)
    col4.metric("üé® Couleurs configur√©es", "5")
    
    # Graphiques par type de label
    st.markdown("### üìä R√©partition des labels par type")
    
    # Compter les labels par type
    label_counts = {'h1': 0, 'h2': 0, 'v1': 0, 'v2': 0}
    
    for zone in st.session_state.zones:
        for label in zone.get('labels', []):
            label_type = label.get('type', '')
            if label_type in label_counts:
                label_counts[label_type] += 1
    
    # Cr√©er le graphique
    col1, col2 = st.columns(2)
    
    with col1:
        # Graphique pour les headers horizontaux
        h_data = pd.DataFrame({
            'Type': ['H1', 'H2'],
            'Nombre': [label_counts['h1'], label_counts['h2']],
            'Couleur': [st.session_state.color_palette['h1_name'], st.session_state.color_palette['h2_name']]
        })
        
        fig_h = px.bar(h_data, x='Type', y='Nombre', 
                      title="Headers Horizontaux",
                      color='Type',
                      color_discrete_map={
                          'H1': f"#{st.session_state.color_palette['h1_color']}",
                          'H2': f"#{st.session_state.color_palette['h2_color']}"
                      })
        st.plotly_chart(fig_h, use_container_width=True)
    
    with col2:
        # Graphique pour les headers verticaux
        v_data = pd.DataFrame({
            'Type': ['V1', 'V2'],
            'Nombre': [label_counts['v1'], label_counts['v2']],
            'Couleur': [st.session_state.color_palette['v1_name'], st.session_state.color_palette['v2_name']]
        })
        
        fig_v = px.bar(v_data, x='Type', y='Nombre',
                      title="Headers Verticaux",
                      color='Type',
                      color_discrete_map={
                          'V1': f"#{st.session_state.color_palette['v1_color']}",
                          'V2': f"#{st.session_state.color_palette['v2_color']}"
                      })
        st.plotly_chart(fig_v, use_container_width=True)
    
    # Tableau r√©capitulatif des zones
    st.markdown("### üìã D√©tail par zone")
    
    zone_data = []
    for zone in st.session_state.zones:
        # Compter les labels par type pour cette zone
        zone_label_counts = {'h1': 0, 'h2': 0, 'v1': 0, 'v2': 0}
        for label in zone.get('labels', []):
            label_type = label.get('type', '')
            if label_type in zone_label_counts:
                zone_label_counts[label_type] += 1
        
        zone_data.append({
            'Zone': zone['id'],
            'Cellules': zone['cell_count'],
            'Position': f"{num_to_excel_col(zone['bounds']['min_col'])}{zone['bounds']['min_row']} - {num_to_excel_col(zone['bounds']['max_col'])}{zone['bounds']['max_row']}",
            'H1': zone_label_counts['h1'],
            'H2': zone_label_counts['h2'],
            'V1': zone_label_counts['v1'],
            'V2': zone_label_counts['v2'],
            'Total Labels': sum(zone_label_counts.values())
        })
    
    st.dataframe(pd.DataFrame(zone_data), use_container_width=True)
    
def load_workbook_with_values(file):
    """
    Charge un fichier Excel avec les valeurs calcul√©es (pas les formules)
    """
    import openpyxl
    import xlrd
    import tempfile
    import os
    
    # D√©terminer le type de fichier
    filename = file.name.lower()
    
    if filename.endswith('.xlsx'):
        # Fichier .xlsx - utiliser openpyxl avec data_only=True
        return openpyxl.load_workbook(file, data_only=True)
    
    elif filename.endswith('.xls'):
        # Fichier .xls - xlrd retourne d√©j√† les valeurs calcul√©es
        return convert_xls_to_openpyxl_values(file)
    
    else:
        raise ValueError("Format de fichier non support√©. Utilisez .xlsx ou .xls")

def convert_xls_to_openpyxl_values(file):
    """
    Convertit un fichier .xls en workbook openpyxl avec les valeurs
    """
    import xlrd
    import openpyxl
    
    # Lire le fichier .xls avec xlrd
    xls_book = xlrd.open_workbook(file_contents=file.read(), formatting_info=True)
    
    # Cr√©er un nouveau workbook openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par d√©faut
    
    # Obtenir les informations de formatage
    xf_list = xls_book.format_map
    
    # Parcourir toutes les feuilles
    for sheet_idx, sheet_name in enumerate(xls_book.sheet_names()):
        xls_sheet = xls_book.sheet_by_name(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        
        # Copier les donn√©es et le formatage
        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell = xls_sheet.cell(row_idx, col_idx)
                
                # √âcrire la valeur (xlrd retourne d√©j√† les valeurs calcul√©es)
                ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell.value)
                
                # Appliquer le formatage si disponible
                if cell.xf_index is not None and cell.xf_index < len(xf_list):
                    xf = xf_list[cell.xf_index]
                    
                    # Extraire la couleur de fond si elle existe
                    if xf.background and hasattr(xf.background, 'pattern_colour_index'):
                        color_idx = xf.background.pattern_colour_index
                        if color_idx and color_idx < len(xls_book.colour_map):
                            rgb = xls_book.colour_map.get(color_idx)
                            if rgb:
                                # Convertir RGB en hex
                                hex_color = '%02x%02x%02x' % rgb[:3]
                                from openpyxl.styles import PatternFill
                                fill = PatternFill(start_color=hex_color, 
                                                 end_color=hex_color, 
                                                 fill_type="solid")
                                ws.cell(row=row_idx + 1, column=col_idx + 1).fill = fill
    
    return wb

def display_instructions():
    """Affiche les instructions d'utilisation pour le syst√®me 4 couleurs"""
    with st.expander("‚ÑπÔ∏è Guide d'utilisation - Syst√®me 4 couleurs"):
        st.markdown("""
        ## üöÄ Comment utiliser l'application
        
        ### üìã Concept du syst√®me √† 4 couleurs
        
        Cette application d√©tecte automatiquement les zones de donn√©es et leurs labels associ√©s en utilisant 5 couleurs :
        - 1 couleur pour les zones de donn√©es
        - 2 couleurs pour les headers horizontaux (H1, H2)
        - 2 couleurs pour les headers verticaux (V1, V2)
        
        ### 1. Structure attendue
        
        ```
        [H1] [H1] [H2] [H2] [H1]  <- Headers horizontaux
        [V1] [Z]  [Z]  [Z]  [Z]   <- V1/V2: Headers verticaux, Z: Zones
        [V2] [Z]  [Z]  [Z]  [Z]
        [V1] [Z]  [Z]  [Z]  [Z]
        ```
        
        ### 2. Logique de d√©tection
        
        Pour chaque cellule de zone :
        - **Verticalement** : Remonte jusqu'au premier header H trouv√©, collecte tous les headers de CETTE couleur, s'arr√™te √† l'autre couleur H
        - **Horizontalement** : Recule jusqu'au premier header V trouv√©, collecte tous les headers de CETTE couleur, s'arr√™te √† l'autre couleur V
        
        ### 3. √âtapes d'utilisation
        
        1. **Charger** votre fichier Excel
        2. **Analyser** les couleurs pr√©sentes dans tout le fichier
        3. **Configurer** les 5 couleurs (zone + 2H + 2V)
        4. **Traiter** les feuilles (individuellement ou toutes)
        5. **Exporter** les r√©sultats en JSON
        
        ### 4. Export
        
        Le JSON export√© contiendra pour chaque cellule de zone :
        - Sa valeur et sa position
        - Tous les headers H de la couleur collect√©e
        - Tous les headers V de la couleur collect√©e
        - Structure optimis√©e pour l'extraction par LLM
        """)

# Fonctions auxiliaires pour l'affichage adapt√© aux paires

def create_excel_visualization_pairs(workbook, sheet_name, zones, selected_zone, color_palette):
    """Cr√©e une visualisation adapt√©e aux paires de labels"""
    # Utiliser la fonction de base en adaptant le mapping des couleurs
    adapted_palette = {
        'zone_color': color_palette['zone_color'],
        'zone_name': color_palette['zone_name'],
        'label_colors': {}
    }
    
    # Convertir les paires en format compatible
    # IMPORTANT: Mapper les types de labels correctement
    for i, pair in enumerate(color_palette.get('label_pairs', [])):
        # Les labels ont des types comme 'h_pair_0', 'v_pair_0' dans zone_detector
        adapted_palette['label_colors'][f'h_pair_{i}'] = pair['horizontal']
        adapted_palette['label_colors'][f'v_pair_{i}'] = pair['vertical']
    
    # Debug
    print(f"DEBUG visualization: adapted_palette = {adapted_palette}")
    
    # V√©rifier si on a des labels dans les zones
    total_labels = sum(len(z.get('labels', [])) for z in zones)
    print(f"DEBUG visualization: Total labels in zones = {total_labels}")
    
    return create_excel_visualization(workbook, sheet_name, zones, selected_zone, adapted_palette)

def create_zone_detail_view_pairs(workbook, sheet_name, zone, color_palette):
    """Cr√©e une vue d√©taill√©e adapt√©e aux paires"""
    # Adapter le format pour r√©utiliser la fonction existante
    adapted_palette = {
        'zone_color': color_palette['zone_color'],
        'zone_name': color_palette['zone_name'],
        'label_colors': {}
    }
    
    # Cr√©er un mapping pour toutes les couleurs de labels
    for i, pair in enumerate(color_palette.get('label_pairs', [])):
        adapted_palette['label_colors'][f'h_pair_{i}'] = pair['horizontal']
        adapted_palette['label_colors'][f'v_pair_{i}'] = pair['vertical']
    
    return create_zone_detail_view(workbook, sheet_name, zone, adapted_palette)

def create_dataframe_view_pairs(workbook, sheet_name: str, zones: List[Dict] = None, 
                               color_palette: Optional[Dict] = None, max_rows: int = 50) -> pd.DataFrame:
    """
    Cr√©e une vue DataFrame styl√©e de la feuille Excel avec coloration des zones et paires
    """
    from utils.excel_utils import num_to_excel_col, excel_col_to_num
    from utils.color_detector import hex_to_rgb
    
    ws = workbook[sheet_name]
    
    # Limiter les dimensions pour la performance
    max_row = min(ws.max_row, max_rows)
    max_col = min(ws.max_column, 26)
    
    # Cr√©er un mapping des cellules color√©es
    colored_cells = {}
    if zones and color_palette:
        for zone in zones:
            # Cellules de la zone
            for cell in zone['cells']:
                if cell['row'] <= max_row and cell['col'] <= max_col:
                    colored_cells[(cell['row'], cell['col'])] = {
                        'color': color_palette['zone_color'],
                        'type': 'zone',
                        'zone_id': zone['id']
                    }
            
            # Labels de la zone avec gestion des paires
            for label in zone.get('labels', []):
                if label['row'] <= max_row and label['col'] <= max_col:
                    # D√©terminer la couleur du label selon la paire
                    label_color = '#888888'  # Couleur par d√©faut
                    
                    if 'pair_id' in label and label['pair_id'] < len(color_palette.get('label_pairs', [])):
                        pair = color_palette['label_pairs'][label['pair_id']]
                        if label.get('direction') == 'horizontal':
                            label_color = pair['horizontal']['color']
                        else:
                            label_color = pair['vertical']['color']
                    
                    colored_cells[(label['row'], label['col'])] = {
                        'color': label_color,
                        'type': 'label',
                        'label_type': f"pair_{label.get('pair_id', 0)}_{label.get('direction', 'unknown')}",
                        'zone_id': zone['id'],
                        'pair_id': label.get('pair_id', 0),
                        'direction': label.get('direction', 'unknown')
                    }
    
    # Cr√©er les donn√©es du DataFrame
    data = []
    columns = [num_to_excel_col(i) for i in range(1, max_col + 1)]
    
    for row in range(1, max_row + 1):
        row_data = []
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value if cell.value is not None else ""
            row_data.append(str(value))
        data.append(row_data)
    
    # Cr√©er le DataFrame
    df = pd.DataFrame(data, columns=columns, index=range(1, max_row + 1))
    
    # Si pas de zones ou de mapping de couleurs, retourner le DataFrame simple
    if not zones or not color_palette:
        return df
    
    # Appliquer le style avec les couleurs
    def style_cells(val):
        """Fonction pour styler les cellules"""
        from utils.excel_utils import excel_col_to_num
        from utils.color_detector import hex_to_rgb
        
        styles = pd.DataFrame('', index=df.index, columns=df.columns)
        
        for row_idx, row_num in enumerate(df.index, 1):
            for col_idx, col_name in enumerate(df.columns):
                col_num = excel_col_to_num(col_name)
                
                if (row_num, col_num) in colored_cells:
                    cell_info = colored_cells[(row_num, col_num)]
                    color = cell_info['color']
                    
                    # Calculer une couleur de texte contrastante
                    r, g, b = hex_to_rgb(color)
                    brightness = (r * 299 + g * 587 + b * 114) / 1000
                    text_color = 'white' if brightness < 128 else 'black'
                    
                    if cell_info['type'] == 'zone':
                        styles.iloc[row_idx-1, col_idx] = f'background-color: #{color}; color: {text_color}; border: 2px solid #{color};'
                    elif cell_info['type'] == 'label':
                        # Style diff√©renci√© selon la direction
                        if cell_info.get('direction') == 'horizontal':
                            styles.iloc[row_idx-1, col_idx] = f'background-color: #{color}; color: {text_color}; border: 3px solid #{color}; font-weight: bold; text-decoration: underline;'
                        else:
                            styles.iloc[row_idx-1, col_idx] = f'background-color: #{color}; color: {text_color}; border: 3px double #{color}; font-weight: bold; font-style: italic;'
        
        return styles
    
    # Appliquer le style
    try:
        styled_df = df.style.apply(style_cells, axis=None)
        styled_df = styled_df.set_table_attributes('style="border-collapse: collapse; font-size: 12px;"')
        return styled_df
    except Exception as e:
        # En cas d'erreur avec le style, retourner le DataFrame simple
        print(f"Erreur lors de l'application du style: {e}")
        return df

def create_zone_detail_table_view_pairs(workbook, sheet_name: str, zone: Dict, color_palette: Dict) -> pd.DataFrame:
    """
    Cr√©e une vue tableau d√©taill√©e pour une zone sp√©cifique avec coloration des paires
    """
    ws = workbook[sheet_name]
    bounds = zone['bounds']
    
    # Ajouter une marge pour voir les labels autour
    margin = 3
    min_row = max(1, bounds['min_row'] - margin)
    max_row = min(ws.max_row, bounds['max_row'] + margin)
    min_col = max(1, bounds['min_col'] - margin)
    max_col = min(ws.max_column, bounds['max_col'] + margin)
    
    # Cr√©er un mapping des cellules de la zone et des labels
    zone_cells = {(c['row'], c['col']) for c in zone['cells']}
    label_cells = {(l['row'], l['col']): l for l in zone.get('labels', [])}
    
    # Cr√©er les donn√©es du DataFrame
    data = []
    columns = [num_to_excel_col(i) for i in range(min_col, max_col + 1)]
    
    for row in range(min_row, max_row + 1):
        row_data = []
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value if cell.value is not None else ""
            row_data.append(str(value))
        data.append(row_data)
    
    # Cr√©er le DataFrame
    df = pd.DataFrame(data, columns=columns, index=range(min_row, max_row + 1))
    
    def style_zone_cells(val):
        """Fonction pour styler les cellules de la zone avec paires"""
        from utils.excel_utils import excel_col_to_num, num_to_excel_col
        from utils.color_detector import hex_to_rgb
        
        styles = pd.DataFrame('', index=df.index, columns=df.columns)
        
        for row_idx in range(len(df)):
            actual_row = df.index[row_idx]
            
            for col_idx in range(len(df.columns)):
                col_name = df.columns[col_idx]
                col_num = excel_col_to_num(col_name)
                
                # V√©rifier si c'est une cellule de zone
                if (actual_row, col_num) in zone_cells:
                    zone_color = color_palette['zone_color']
                    r, g, b = hex_to_rgb(zone_color)
                    brightness = (r * 299 + g * 587 + b * 114) / 1000
                    text_color = 'white' if brightness < 128 else 'black'
                    
                    styles.iloc[row_idx, col_idx] = f'background-color: #{zone_color}; color: {text_color}; font-weight: bold; border: 2px solid #{zone_color};'
                
                # V√©rifier si c'est un label (priorit√© sur la zone)
                elif (actual_row, col_num) in label_cells:
                    label = label_cells[(actual_row, col_num)]
                    
                    # D√©terminer la couleur du label selon la paire et la direction
                    label_color = None
                    if 'pair_id' in label and label['pair_id'] < len(color_palette.get('label_pairs', [])):
                        pair = color_palette['label_pairs'][label['pair_id']]
                        if label.get('direction') == 'horizontal':
                            label_color = pair['horizontal']['color']
                        else:
                            label_color = pair['vertical']['color']
                    
                    if label_color:
                        r, g, b = hex_to_rgb(label_color)
                        brightness = (r * 299 + g * 587 + b * 114) / 1000
                        text_color = 'white' if brightness < 128 else 'black'
                        
                        # Style diff√©renci√© selon la direction
                        if label.get('direction') == 'horizontal':
                            styles.iloc[row_idx, col_idx] = f'background-color: #{label_color}; color: {text_color}; font-weight: bold; border: 3px solid #{label_color}; box-shadow: 0 2px 0 rgba({r},{g},{b},0.7);'
                        else:
                            styles.iloc[row_idx, col_idx] = f'background-color: #{label_color}; color: {text_color}; font-weight: bold; border: 3px solid #{label_color}; box-shadow: 2px 0 0 rgba({r},{g},{b},0.7);'
        
        return styles

    # Appliquer le style
    try:
        styled_df = df.style.apply(style_zone_cells, axis=None)
        styled_df = styled_df.set_table_attributes('style="border-collapse: collapse; font-size: 14px;"')
        styled_df = styled_df.set_caption(f"Zone {zone['id']} - Vue d√©taill√©e tableau (Paires de labels)")
        return styled_df
    except Exception as e:
        print(f"Erreur lors de l'application du style: {e}")
        return df

def create_zone_detail_table_view_pairs_enhanced(workbook, sheet_name: str, zone: Dict, 
                                                color_palette: Dict, show_markers: bool = True) -> pd.DataFrame:
    """
    Version am√©lior√©e de la vue tableau avec marqueurs visuels pour les paires
    """
    ws = workbook[sheet_name]
    bounds = zone['bounds']
    
    # Calculer la zone d'affichage avec marge
    margin = 3
    min_row = max(1, bounds['min_row'] - margin)
    max_row = min(ws.max_row, bounds['max_row'] + margin)
    min_col = max(1, bounds['min_col'] - margin)
    max_col = min(ws.max_column, bounds['max_col'] + margin)
    
    # Cr√©er les mappings
    zone_cells = {(c['row'], c['col']) for c in zone['cells']}
    label_cells = {(l['row'], l['col']): l for l in zone.get('labels', [])}
    
    # Cr√©er le DataFrame avec marqueurs visuels
    columns = [num_to_excel_col(i) for i in range(min_col, max_col + 1)]
    data = []
    
    for row in range(min_row, max_row + 1):
        row_data = []
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value if cell.value is not None else ""
            
            # Ajouter des indicateurs visuels dans le texte si activ√©
            if show_markers:
                if (row, col) in zone_cells:
                    # Cellule de zone
                    value = f"üîµ {value}" if value else "üîµ"
                elif (row, col) in label_cells:
                    # Label - indicateur selon la paire et la direction
                    label = label_cells[(row, col)]
                    if 'pair_id' in label:
                        pair_num = label['pair_id'] + 1
                        if label.get('direction') == 'horizontal':
                            marker = f"‚û°Ô∏èP{pair_num}"
                        else:
                            marker = f"‚¨áÔ∏èP{pair_num}"
                        value = f"{marker} {value}" if value else marker
            
            row_data.append(str(value))
        data.append(row_data)
    
    df = pd.DataFrame(data, columns=columns, index=range(min_row, max_row + 1))
    
    # Style avanc√© avec CSS
    def enhanced_style(x):
        """Style avanc√© pour le tableau avec paires"""
        from utils.excel_utils import excel_col_to_num
        from utils.color_detector import hex_to_rgb
        
        styles = pd.DataFrame('', index=df.index, columns=df.columns)
        
        for row_idx in range(len(df)):
            actual_row = df.index[row_idx]
            
            for col_idx in range(len(df.columns)):
                col_name = df.columns[col_idx]
                col_num = excel_col_to_num(col_name)
                
                if (actual_row, col_num) in zone_cells:
                    # Style pour cellules de zone
                    zone_color = color_palette['zone_color']
                    r, g, b = hex_to_rgb(zone_color)
                    
                    styles.iloc[row_idx, col_idx] = f'background-color: rgba({r}, {g}, {b}, 0.3); border: 3px solid #{zone_color}; font-weight: bold; text-align: center;'
                
                elif (actual_row, col_num) in label_cells:
                    # Style pour labels avec diff√©renciation par paire
                    label = label_cells[(actual_row, col_num)]
                    
                    if 'pair_id' in label and label['pair_id'] < len(color_palette.get('label_pairs', [])):
                        pair = color_palette['label_pairs'][label['pair_id']]
                        
                        if label.get('direction') == 'horizontal':
                            label_color = pair['horizontal']['color']
                            r, g, b = hex_to_rgb(label_color)
                            styles.iloc[row_idx, col_idx] = f'background-color: rgba({r}, {g}, {b}, 0.5); border-top: 4px solid #{label_color}; border-bottom: 4px solid #{label_color}; font-weight: bold; text-align: center;'
                        else:
                            label_color = pair['vertical']['color']
                            r, g, b = hex_to_rgb(label_color)
                            styles.iloc[row_idx, col_idx] = f'background-color: rgba({r}, {g}, {b}, 0.5); border-left: 4px solid #{label_color}; border-right: 4px solid #{label_color}; font-weight: bold; text-align: center;'
        
        return styles
    
    try:
        styled_df = df.style.apply(enhanced_style, axis=None)
        styled_df = styled_df.set_table_attributes('style="border-collapse: collapse; font-size: 12px;"')
        
        # Cr√©er la l√©gende
        legend_parts = ["üîµ = Zone"]
        for i, pair in enumerate(color_palette.get('label_pairs', [])):
            legend_parts.append(f"‚û°Ô∏èP{i+1} = {pair['horizontal']['name']}")
            legend_parts.append(f"‚¨áÔ∏èP{i+1} = {pair['vertical']['name']}")
        
        caption = f"<h3>Zone {zone['id']} - Vue avec marqueurs de paires</h3><p>{' | '.join(legend_parts)}</p>"
        styled_df = styled_df.set_caption(caption)
        
        return styled_df
    except Exception as e:
        print(f"Erreur style avanc√©: {e}")
        return df

def export_to_json_pairs(zones, sheet_name, color_palette):
    """Exporte les zones avec le syst√®me de 4 couleurs en JSON"""
    import json
    from datetime import datetime
    from utils.excel_utils import num_to_excel_col
    
    export_data = {
        "date_export": datetime.now().isoformat(),
        "sheet_name": sheet_name,
        "detection_mode": "two_colors_system",
        "color_palette": {
            "zone_color": f"#{color_palette['zone_color']}",
            "zone_name": color_palette['zone_name'],
            "headers": {
                "horizontal": {
                    "h1": {
                        "color": f"#{color_palette.get('h1_color', '')}",
                        "name": color_palette.get('h1_name', 'H1')
                    },
                    "h2": {
                        "color": f"#{color_palette.get('h2_color', '')}",
                        "name": color_palette.get('h2_name', 'H2')
                    }
                },
                "vertical": {
                    "v1": {
                        "color": f"#{color_palette.get('v1_color', '')}",
                        "name": color_palette.get('v1_name', 'V1')
                    },
                    "v2": {
                        "color": f"#{color_palette.get('v2_color', '')}",
                        "name": color_palette.get('v2_name', 'V2')
                    }
                }
            }
        },
        "zones": []
    }
    
    # Exporter les zones
    for zone in zones:
        zone_data = {
            "id": zone['id'],
            "bounds": {
                "min_row": zone['bounds']['min_row'],
                "max_row": zone['bounds']['max_row'],
                "min_col": zone['bounds']['min_col'],
                "max_col": zone['bounds']['max_col'],
                "min_col_letter": num_to_excel_col(zone['bounds']['min_col']),
                "max_col_letter": num_to_excel_col(zone['bounds']['max_col'])
            },
            "cell_count": zone['cell_count'],
            "cells": format_cells_for_export_pairs(zone['cells']),
            "labels": {
                "horizontal": [],
                "vertical": []
            }
        }
        
        # Organiser les labels par direction
        for label in zone.get('labels', []):
            formatted_label = {
                "address": f"{num_to_excel_col(label['col'])}{label['row']}",
                "row": label['row'],
                "col": label['col'],
                "col_letter": num_to_excel_col(label['col']),
                "value": str(label.get('value', '')) if label.get('value') is not None else "",
                "type": label.get('type', ''),
                "distance": label.get('distance', 0),
                "color": f"#{label.get('color', '')}"
            }
            
            if label.get('direction') == 'horizontal':
                zone_data["labels"]["horizontal"].append(formatted_label)
            else:
                zone_data["labels"]["vertical"].append(formatted_label)
        
        export_data["zones"].append(zone_data)
    
    return json.dumps(export_data, indent=2, ensure_ascii=False)

def format_cells_for_export_pairs(cells):
    """Formate les cellules pour l'export"""
    from utils.excel_utils import num_to_excel_col
    formatted_cells = []
    
    for cell in cells:
        formatted_cells.append({
            "address": f"{num_to_excel_col(cell['col'])}{cell['row']}",
            "row": cell['row'],
            "col": cell['col'],
            "col_letter": num_to_excel_col(cell['col']),
            "value": str(cell.get('value', '')) if cell.get('value') is not None else ""
        })
    
    return formatted_cells

def format_labels_by_pair(labels):
    """Organise les labels par paire pour l'export"""
    from utils.excel_utils import num_to_excel_col
    from collections import defaultdict
    
    labels_by_pair = defaultdict(lambda: {"horizontal": [], "vertical": []})
    
    for label in labels:
        if 'pair_id' in label:
            formatted_label = {
                "address": f"{num_to_excel_col(label['col'])}{label['row']}",
                "row": label['row'],
                "col": label['col'],
                "col_letter": num_to_excel_col(label['col']),
                "value": str(label.get('value', '')) if label.get('value') is not None else "",
                "distance": label.get('distance', 0),
                "position": label.get('position', ''),
                "pair_name": label.get('pair_name', '')
            }
            
            labels_by_pair[label['pair_id']][label['direction']].append(formatted_label)
    
    # Convertir en format liste pour JSON
    result = []
    for pair_id in sorted(labels_by_pair.keys()):
        result.append({
            "pair_id": pair_id,
            "horizontal_labels": labels_by_pair[pair_id]["horizontal"],
            "vertical_labels": labels_by_pair[pair_id]["vertical"]
        })
    
    return result

if __name__ == "__main__":
    main()